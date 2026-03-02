# excel_processor.py
# ─────────────────────────────────────────────────────────────
#  Handles all reading from and writing back to the Excel file.
# ─────────────────────────────────────────────────────────────

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from copy import copy


# ── Column names exactly as they appear in the Excel ──────────
COL_SESSION  = "SESSION_ID"
COL_QUERY    = "QUERY"
COL_RESPONSE = "RESPONSE"
COL_LLM_GCR  = "LLM GCR"
COL_YES_NO   = "LLM GCR yes/no"


def load_excel(file_path: str) -> pd.DataFrame:
    """
    Load the Excel file into a DataFrame.
    Keeps original row order via the internal pandas index.
    """
    df = pd.read_excel(file_path, dtype=str)   # read everything as string to avoid type coercion
    df = df.fillna("")                          # replace NaN with empty string
    return df


def group_sessions(df: pd.DataFrame) -> dict:
    """
    Group rows by SESSION_ID.
    Returns:
        {
          session_id: [
              {"row_index": int, "query": str, "response": str},
              ...
          ]
        }
    Preserves original row order within each session.
    """
    sessions = {}
    for idx, row in df.iterrows():
        sid = row[COL_SESSION].strip()
        if not sid:
            continue  # skip rows with no session id

        if sid not in sessions:
            sessions[sid] = []

        sessions[sid].append({
            "row_index": idx,
            "query":     row[COL_QUERY].strip(),
            "response":  row[COL_RESPONSE].strip(),
        })

    return sessions


def write_results_to_excel(
    source_path: str,
    output_path: str,
    results: dict
) -> None:
    """
    Write evaluation results back to Excel without destroying
    existing formatting (colours, borders, fonts, column widths).

    results format:
        {
          session_id: {
              "llm_gcr":   "Full Gemini output text",
              "yes_no":    "Yes" | "No" | "ERROR",
              "row_indices": [list of int row indices for this session]
          }
        }

    Strategy:
    - openpyxl loads the workbook preserving all formatting.
    - We only write to COL_LLM_GCR and COL_YES_NO columns.
    - For multi-row sessions, the LLM GCR text is written only on
      the FIRST row of the session; remaining rows in that session
      get "↑ same session" as a marker so the file stays readable.
    - The yes/no value is written on EVERY row of the session so
      filters work correctly in Excel.
    """
    wb = openpyxl.load_workbook(source_path)
    ws = wb.active

    # ── Locate column letters from header row (row 1) ─────────
    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[str(cell.value).strip()] = cell.column

    gcr_col  = header_map.get(COL_LLM_GCR)
    yesno_col = header_map.get(COL_YES_NO)

    if not gcr_col or not yesno_col:
        raise ValueError(
            f"Could not find columns '{COL_LLM_GCR}' or '{COL_YES_NO}' "
            f"in the Excel header. Found: {list(header_map.keys())}"
        )

    # ── Colour fills for visual feedback ──────────────────────
    green_fill  = PatternFill("solid", fgColor="C6EFCE")  # light green  → Yes
    red_fill    = PatternFill("solid", fgColor="FFC7CE")  # light red    → No
    orange_fill = PatternFill("solid", fgColor="FFEB9C")  # light orange → ERROR

    # Build a quick lookup: row_index (0-based df) → result
    row_to_session = {}
    for sid, data in results.items():
        for i, ridx in enumerate(data["row_indices"]):
            row_to_session[ridx] = {
                "session_id": sid,
                "llm_gcr":   data["llm_gcr"] if i == 0 else "↑ see first row of this session",
                "yes_no":    data["yes_no"],
                "is_first":  i == 0,
            }

    # Excel rows are 1-indexed; row 1 = header → data starts at row 2
    for df_idx, info in row_to_session.items():
        excel_row = df_idx + 2   # +1 for header, +1 for 0→1 index

        yes_no_val = info["yes_no"]

        # ── Write LLM GCR cell ────────────────────────────────
        gcr_cell = ws.cell(row=excel_row, column=gcr_col)
        gcr_cell.value     = info["llm_gcr"]
        gcr_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # ── Write Yes/No cell ─────────────────────────────────
        yn_cell = ws.cell(row=excel_row, column=yesno_col)
        yn_cell.value     = yes_no_val
        yn_cell.alignment = Alignment(horizontal="center", vertical="center")
        yn_cell.font      = Font(bold=True)

        # Apply colour based on value
        if yes_no_val == "Yes":
            yn_cell.fill  = green_fill
        elif yes_no_val == "No":
            yn_cell.fill  = red_fill
        else:
            yn_cell.fill  = orange_fill   # ERROR or PARSE_ERROR

    wb.save(output_path)
    print(f"[excel_processor] Results written to: {output_path}")